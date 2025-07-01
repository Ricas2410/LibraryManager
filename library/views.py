from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse
from datetime import timedelta
import json
from functools import wraps

from .models import (User, Book, Author, Category, Publisher, Loan, Reservation,
                     LibrarySettings, AuditLog, Section, ShelfLocation, Floor, Notification)
from .forms import (CustomUserCreationForm, CustomAuthenticationForm, UserProfileForm,
                   BookForm, BookSearchForm, LoanForm, ReservationForm, LibrarySettingsForm)


# Custom decorator for API endpoints
def api_login_required(view_func):
    """
    Decorator for API views that returns JSON error instead of redirecting to login
    """
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return JsonResponse({
                'error': 'Authentication required',
                'notifications': [],
                'unread_count': 0
            }, status=401)
        return view_func(request, *args, **kwargs)
    return wrapper


def is_admin(user):
    """Check if user is admin"""
    return user.is_authenticated and user.role == 'admin'


def is_librarian_or_admin(user):
    """Check if user is librarian or admin"""
    return user.is_authenticated and user.role in ['admin', 'librarian']


def can_manage_books(user):
    """Check if user can manage books"""
    return user.is_authenticated and user.can_manage_books()


def home(request):
    """Home page with dashboard"""
    if not request.user.is_authenticated:
        return redirect('login')

    context = {
        'total_books': Book.objects.count(),
        'available_books': Book.objects.filter(status='available').count(),
        'total_users': User.objects.filter(is_active_member=True).count(),
        'active_loans': Loan.objects.filter(status='active').count(),
        'overdue_loans': Loan.objects.filter(status='overdue').count(),
    }

    # Recent activities for admins/librarians
    if request.user.can_manage_books():
        context['recent_loans'] = Loan.objects.select_related('book', 'borrower').order_by('-issue_date')[:5]
        context['recent_reservations'] = Reservation.objects.select_related('book', 'user').filter(status='active').order_by('-reservation_date')[:5]

        # Add admin notification counts
        from datetime import timedelta

        context['pending_requests_count'] = Loan.objects.filter(status='pending').count()
        context['active_reservations_count'] = Reservation.objects.filter(status='active').count()
        context['overdue_loans_count'] = Loan.objects.filter(
            status='active',
            due_date__lt=timezone.now()
        ).count()

        # Count books due soon (next 3 days)
        today = timezone.now().date()
        due_soon_date = today + timedelta(days=3)
        context['due_soon_loans_count'] = Loan.objects.filter(
            status='active',
            due_date__date__gte=today,
            due_date__date__lte=due_soon_date
        ).count()

    # User's personal data
    if request.user.can_borrow_books():
        context['my_loans'] = Loan.objects.filter(borrower=request.user, status='active').select_related('book')
        context['my_reservations'] = Reservation.objects.filter(user=request.user, status='active').select_related('book')

    return render(request, 'library/home.html', context)


@login_required
def student_dashboard(request):
    """Student dashboard with borrowing functionality"""
    if not request.user.can_borrow_books():
        messages.error(request, 'Access denied. Students and teachers only.')
        return redirect('home')

    # Get user's current loans
    active_loans = Loan.objects.filter(
        borrower=request.user,
        status='active'
    ).select_related('book').order_by('-issue_date')

    # Get user's pending requests
    pending_requests = Loan.objects.filter(
        borrower=request.user,
        status='pending'
    ).select_related('book').order_by('-issue_date')

    # Get user's loan history
    loan_history = Loan.objects.filter(
        borrower=request.user
    ).select_related('book').order_by('-issue_date')[:10]

    # Get user's reservations
    reservations = Reservation.objects.filter(
        user=request.user,
        status='active'
    ).select_related('book').order_by('priority')

    # Get overdue loans
    overdue_loans = Loan.objects.filter(
        borrower=request.user,
        status='overdue'
    ).select_related('book')

    # Get available books for borrowing
    available_books = Book.objects.filter(
        status='available',
        available_copies__gt=0
    ).select_related('publisher').prefetch_related('authors')[:12]

    # Calculate statistics
    stats = {
        'total_loans': loan_history.count(),
        'active_loans': active_loans.count(),
        'overdue_loans': overdue_loans.count(),
        'reservations': reservations.count(),
        'pending_requests': pending_requests.count(),
    }

    context = {
        'active_loans': active_loans,
        'loan_history': loan_history,
        'reservations': reservations,
        'overdue_loans': overdue_loans,
        'available_books': available_books,
        'pending_requests': pending_requests,
        'stats': stats,
    }

    return render(request, 'library/student_dashboard.html', context)


@login_required
def request_book(request, book_id):
    """Request a book for borrowing"""
    if not request.user.can_borrow_books():
        messages.error(request, 'You do not have permission to borrow books.')
        return redirect('book_detail', book_id=book_id)

    book = get_object_or_404(Book, id=book_id)

    # Check if book is available
    if not book.is_available():
        messages.error(request, 'This book is currently not available.')
        return redirect('book_detail', book_id=book_id)

    # Check if user already has this book (active loan or pending request)
    existing_loan = Loan.objects.filter(
        borrower=request.user,
        book=book,
        status__in=['active', 'pending']
    ).exists()

    if existing_loan:
        # Check if it's active or pending
        loan_status = Loan.objects.filter(
            borrower=request.user,
            book=book,
            status__in=['active', 'pending']
        ).first().status

        if loan_status == 'active':
            messages.error(request, 'You already have this book on loan.')
        else:
            messages.error(request, 'You already have a pending request for this book.')
        return redirect('book_detail', book_id=book_id)

    # Check loan limits
    active_loans_count = Loan.objects.filter(
        borrower=request.user,
        status='active'
    ).count()

    # Get library settings for max books per user
    try:
        settings = LibrarySettings.objects.first()
        max_books = settings.max_books_per_user if settings else 5
    except:
        max_books = 5

    if active_loans_count >= max_books:
        messages.error(request, f'You have reached the maximum limit of {max_books} books.')
        return redirect('book_detail', book_id=book_id)

    if request.method == 'POST':
        print(f"🔍 DEBUG: POST request received for book request - Book: {book.title}, User: {request.user.username}")

        # Create a pending loan request
        try:
            settings = LibrarySettings.objects.first()
            loan_period = settings.default_loan_period if settings else 14
        except:
            loan_period = 14

        due_date = timezone.now() + timedelta(days=loan_period)

        print(f"🔍 DEBUG: Creating loan with due_date: {due_date}")

        loan = Loan.objects.create(
            book=book,
            borrower=request.user,
            issued_by=None,  # Will be set when approved
            due_date=due_date,
            status='pending'  # Pending approval
        )

        print(f"🔍 DEBUG: Loan created with ID: {loan.id}")

        # Log the request
        AuditLog.objects.create(
            action='loan_requested',
            user=request.user,
            target_user=request.user,
            book=book,
            description=f"Book '{book.title}' requested by {request.user.get_full_name()}",
            ip_address=request.META.get('REMOTE_ADDR')
        )

        print(f"🔍 DEBUG: Audit log created, redirecting to student_dashboard")

        messages.success(request, f'Your request for "{book.title}" has been submitted. Please visit the librarian to collect the book.')
        return redirect('student_dashboard')

    context = {
        'book': book,
    }

    return render(request, 'library/request_book.html', context)


def custom_login(request):
    """Custom login view"""
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)

            # Log the login
            AuditLog.objects.create(
                action='user_login',
                user=user,
                description=f"User {user.username} logged in",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            messages.success(request, f'Welcome back, {user.get_full_name()}!')
            next_url = request.GET.get('next', 'home')
            return redirect(next_url)
    else:
        form = CustomAuthenticationForm()

    return render(request, 'registration/login.html', {'form': form})


def custom_logout(request):
    """Custom logout view"""
    if request.user.is_authenticated:
        # Log the logout
        AuditLog.objects.create(
            action='user_logout',
            user=request.user,
            description=f"User {request.user.username} logged out",
            ip_address=request.META.get('REMOTE_ADDR')
        )

    logout(request)
    messages.info(request, 'You have been logged out successfully.')
    return redirect('login')


@user_passes_test(is_admin)
def register_user(request):
    """Register new user (admin only)"""
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST, user=request.user)
        if form.is_valid():
            user = form.save()

            # Log the user creation
            AuditLog.objects.create(
                action='user_created',
                user=request.user,
                target_user=user,
                description=f"User {user.username} created by {request.user.username}",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            messages.success(request, f'User {user.username} created successfully!')
            return redirect('user_list')
    else:
        form = CustomUserCreationForm(user=request.user)

    return render(request, 'library/user_management.html', {'form': form})


@login_required
def profile(request):
    """User profile view"""
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()

            # Log the profile update
            AuditLog.objects.create(
                action='user_updated',
                user=request.user,
                description=f"User {request.user.username} updated their profile",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            messages.success(request, 'Profile updated successfully!')
            return redirect('profile')
    else:
        form = UserProfileForm(instance=request.user)

    # Get user's loan and reservation history
    loans = Loan.objects.filter(borrower=request.user).select_related('book').order_by('-issue_date')
    reservations = Reservation.objects.filter(user=request.user).select_related('book').order_by('-reservation_date')

    context = {
        'form': form,
        'loans': loans[:10],  # Last 10 loans
        'reservations': reservations[:10],  # Last 10 reservations
    }

    return render(request, 'library/profile.html', context)


def book_list(request):
    """List all books with search and filtering"""
    form = BookSearchForm(request.GET)
    books = Book.objects.select_related('publisher').prefetch_related('authors', 'categories')

    if form.is_valid():
        query = form.cleaned_data.get('query')
        category = form.cleaned_data.get('category')
        status = form.cleaned_data.get('status')
        section = form.cleaned_data.get('section')

        if query:
            books = books.filter(
                Q(title__icontains=query) |
                Q(authors__first_name__icontains=query) |
                Q(authors__last_name__icontains=query) |
                Q(isbn__icontains=query) |
                Q(publisher__name__icontains=query)
            ).distinct()

        if category:
            books = books.filter(categories=category)

        if status:
            books = books.filter(status=status)

        if section:
            books = books.filter(section__icontains=section)

    # Pagination
    paginator = Paginator(books, 12)  # 12 books per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'form': form,
        'page_obj': page_obj,
        'books': page_obj,
    }

    return render(request, 'library/book_list.html', context)


def book_detail(request, book_id):
    """Book detail view"""
    book = get_object_or_404(Book, id=book_id)

    # Check if user can reserve this book
    can_reserve = False
    if request.user.is_authenticated and request.user.can_borrow_books():
        can_reserve = (
            not book.is_available() and
            not Reservation.objects.filter(
                book=book,
                user=request.user,
                status='active'
            ).exists()
        )

    # Get active reservations for this book
    reservations = Reservation.objects.filter(
        book=book,
        status='active'
    ).select_related('user').order_by('priority')

    # Get loan history (for librarians/admins)
    loan_history = None
    if request.user.is_authenticated and request.user.can_manage_books():
        loan_history = Loan.objects.filter(book=book).select_related('borrower').order_by('-issue_date')[:10]

    context = {
        'book': book,
        'can_reserve': can_reserve,
        'reservations': reservations,
        'loan_history': loan_history,
    }

    return render(request, 'library/book_detail.html', context)


@user_passes_test(can_manage_books)
def book_add(request):
    """Add new book"""
    if request.method == 'POST':
        form = BookForm(request.POST, request.FILES)
        if form.is_valid():
            book = form.save(commit=False)
            book.created_by = request.user
            book.save()
            form.save_m2m()

            # Log the book addition
            AuditLog.objects.create(
                action='book_added',
                user=request.user,
                book=book,
                description=f"Book '{book.title}' added by {request.user.username}",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            messages.success(request, f'Book "{book.title}" added successfully!')
            return redirect('book_detail', book_id=book.id)
    else:
        form = BookForm()

    return render(request, 'library/book_form.html', {'form': form, 'title': 'Add Book'})


@user_passes_test(can_manage_books)
def book_edit(request, book_id):
    """Edit existing book"""
    book = get_object_or_404(Book, id=book_id)

    if request.method == 'POST':
        form = BookForm(request.POST, request.FILES, instance=book)
        if form.is_valid():
            book = form.save()

            # Log the book update
            AuditLog.objects.create(
                action='book_updated',
                user=request.user,
                book=book,
                description=f"Book '{book.title}' updated by {request.user.username}",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            messages.success(request, f'Book "{book.title}" updated successfully!')
            return redirect('book_detail', book_id=book.id)
    else:
        form = BookForm(instance=book)

    return render(request, 'library/book_form.html', {'form': form, 'title': 'Edit Book', 'book': book})


@user_passes_test(can_manage_books)
def book_delete(request, book_id):
    """Delete book"""
    book = get_object_or_404(Book, id=book_id)

    if request.method == 'POST':
        # Check if book has active loans
        if Loan.objects.filter(book=book, status='active').exists():
            messages.error(request, 'Cannot delete book with active loans.')
            return redirect('book_detail', book_id=book.id)

        book_title = book.title

        # Log the book deletion
        AuditLog.objects.create(
            action='book_deleted',
            user=request.user,
            description=f"Book '{book_title}' deleted by {request.user.username}",
            ip_address=request.META.get('REMOTE_ADDR')
        )

        book.delete()
        messages.success(request, f'Book "{book_title}" deleted successfully!')
        return redirect('book_list')

    return render(request, 'library/book_confirm_delete.html', {'book': book})


@user_passes_test(is_admin)
def user_list(request):
    """List all users (admin only)"""
    users = User.objects.all().order_by('-date_joined')

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(enrollment_number__icontains=search_query)
        )

    # Role filter
    role_filter = request.GET.get('role', '')
    if role_filter:
        users = users.filter(role=role_filter)

    # Pagination
    paginator = Paginator(users, 20)  # 20 users per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'users': page_obj,
        'search_query': search_query,
        'role_filter': role_filter,
        'role_choices': User.ROLE_CHOICES,
    }

    return render(request, 'library/user_list.html', context)


@user_passes_test(is_admin)
def user_detail(request, user_id):
    """User detail view (admin only)"""
    user_obj = get_object_or_404(User, id=user_id)

    # Get user's loan history
    loans = Loan.objects.filter(borrower=user_obj).select_related('book').order_by('-issue_date')

    # Get user's reservation history
    reservations = Reservation.objects.filter(user=user_obj).select_related('book').order_by('-reservation_date')

    # Get statistics
    stats = {
        'total_loans': loans.count(),
        'active_loans': loans.filter(status='active').count(),
        'overdue_loans': loans.filter(status='overdue').count(),
        'total_reservations': reservations.count(),
        'active_reservations': reservations.filter(status='active').count(),
    }

    context = {
        'user_obj': user_obj,
        'loans': loans[:10],  # Last 10 loans
        'reservations': reservations[:10],  # Last 10 reservations
        'stats': stats,
    }

    return render(request, 'library/user_detail.html', context)


@user_passes_test(is_admin)
def user_edit(request, user_id):
    """Edit user (admin only)"""
    user_obj = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=user_obj)
        if form.is_valid():
            form.save()

            # Log the user update
            AuditLog.objects.create(
                action='user_updated',
                user=request.user,
                target_user=user_obj,
                description=f"User {user_obj.username} updated by {request.user.username}",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            messages.success(request, f'User {user_obj.username} updated successfully!')
            return redirect('user_detail', user_id=user_obj.id)
    else:
        form = UserProfileForm(instance=user_obj)

    return render(request, 'library/user_form.html', {'form': form, 'user_obj': user_obj, 'title': 'Edit User'})


@user_passes_test(is_admin)
def user_toggle_status(request, user_id):
    """Toggle user active status (admin only)"""
    user_obj = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        user_obj.is_active_member = not user_obj.is_active_member
        user_obj.save()

        status = "activated" if user_obj.is_active_member else "deactivated"

        # Log the status change
        AuditLog.objects.create(
            action='user_updated',
            user=request.user,
            target_user=user_obj,
            description=f"User {user_obj.username} {status} by {request.user.username}",
            ip_address=request.META.get('REMOTE_ADDR')
        )

        messages.success(request, f'User {user_obj.username} has been {status}.')

    return redirect('user_detail', user_id=user_obj.id)


@user_passes_test(is_librarian_or_admin)
def loan_list(request):
    """List all loans"""
    loans = Loan.objects.select_related('book', 'borrower').order_by('-issue_date')

    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        loans = loans.filter(status=status_filter)

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        loans = loans.filter(
            Q(book__title__icontains=search_query) |
            Q(borrower__username__icontains=search_query) |
            Q(borrower__first_name__icontains=search_query) |
            Q(borrower__last_name__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(loans, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'loans': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'status_choices': Loan.STATUS_CHOICES,
    }

    return render(request, 'library/loan_list.html', context)


@user_passes_test(is_librarian_or_admin)
def loan_create(request):
    """Create new loan"""
    if request.method == 'POST':
        form = LoanForm(request.POST)
        if form.is_valid():
            loan = form.save(commit=False)
            loan.issued_by = request.user

            # Check if book is available
            book = loan.book
            if not book.is_available():
                messages.error(request, 'This book is not available for loan.')
                return render(request, 'library/loan_form.html', {'form': form, 'title': 'Issue Loan'})

            # Update book availability
            book.available_copies -= 1
            book.update_availability()

            loan.save()

            # Log the loan
            AuditLog.objects.create(
                action='loan_issued',
                user=request.user,
                target_user=loan.borrower,
                book=loan.book,
                description=f"Book '{loan.book.title}' issued to {loan.borrower.get_full_name()} by {request.user.username}",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            messages.success(request, f'Book "{loan.book.title}" issued to {loan.borrower.get_full_name()} successfully!')
            return redirect('loan_list')
    else:
        form = LoanForm()

    return render(request, 'library/loan_form.html', {'form': form, 'title': 'Issue Loan'})


@user_passes_test(is_librarian_or_admin)
def loan_return(request, loan_id):
    """Return a book"""
    loan = get_object_or_404(Loan, id=loan_id)

    if request.method == 'POST':
        if loan.status != 'active':
            messages.error(request, 'This loan is not active.')
            return redirect('loan_list')

        # Mark loan as returned
        loan.return_date = timezone.now()
        loan.status = 'returned'
        loan.save()

        # Update book availability
        book = loan.book
        book.available_copies += 1
        book.update_availability()

        # Check for reservations and notify
        next_reservation = Reservation.objects.filter(
            book=book,
            status='active'
        ).order_by('priority').first()

        if next_reservation:
            # Here you could send notification to the user
            messages.info(request, f'Book is reserved for {next_reservation.user.get_full_name()}.')

        # Log the return
        AuditLog.objects.create(
            action='loan_returned',
            user=request.user,
            target_user=loan.borrower,
            book=loan.book,
            description=f"Book '{loan.book.title}' returned by {loan.borrower.get_full_name()}",
            ip_address=request.META.get('REMOTE_ADDR')
        )

        # Create notification for the borrower
        create_notification(
            user=loan.borrower,
            title="Book Returned Successfully",
            message=f"Your book '{loan.book.title}' has been returned successfully. Thank you!",
            notification_type='book_returned',
            book=loan.book,
            loan=loan
        )

        # Send email notification
        send_email_notification(
            user=loan.borrower,
            subject="Book Returned Successfully",
            message=f"""Dear {loan.borrower.get_full_name()},

Your book "{loan.book.title}" has been returned successfully.

Return Details:
- Book: {loan.book.title}
- Return Date: {loan.return_date.strftime('%B %d, %Y at %I:%M %p')}
- Loan Period: {(loan.return_date.date() - loan.issue_date.date()).days} days

Thank you for using the library!

Best regards,
Library Team"""
        )

        messages.success(request, f'Book "{loan.book.title}" returned successfully!')
        return redirect('loan_list')

    return render(request, 'library/loan_return_confirm.html', {'loan': loan})


@user_passes_test(is_librarian_or_admin)
def loan_renew(request, loan_id):
    """Renew a loan via AJAX"""
    from django.http import JsonResponse
    import json

    loan = get_object_or_404(Loan, id=loan_id)

    if request.method == 'POST':
        try:
            if not loan.can_renew():
                return JsonResponse({
                    'success': False,
                    'error': 'This loan cannot be renewed. It may be overdue, already renewed maximum times, or not active.'
                })

            # Renew the loan
            success = loan.renew()

            if success:
                # Log the renewal
                AuditLog.objects.create(
                    action='loan_renewed',
                    user=request.user,
                    target_user=loan.borrower,
                    book=loan.book,
                    description=f"Loan for '{loan.book.title}' renewed by {request.user.username}. New due date: {loan.due_date.strftime('%Y-%m-%d')}",
                    ip_address=request.META.get('REMOTE_ADDR')
                )

                # Create notification for the borrower
                create_notification(
                    user=loan.borrower,
                    title="Loan Renewed",
                    message=f"Your loan for '{loan.book.title}' has been renewed. New due date: {loan.due_date.strftime('%B %d, %Y')}",
                    notification_type='loan_renewed',
                    book=loan.book,
                    loan=loan
                )

                return JsonResponse({
                    'success': True,
                    'new_due_date': loan.due_date.strftime('%b %d, %Y'),
                    'renewal_count': loan.renewal_count,
                    'can_renew': loan.can_renew(),
                    'message': 'Loan renewed successfully!'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Failed to renew loan. Please check loan status and renewal limits.'
                })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error renewing loan: {str(e)}'
            })

    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })


@login_required
def reservation_create(request, book_id):
    """Create a reservation for a book"""
    book = get_object_or_404(Book, id=book_id)

    if not request.user.can_borrow_books():
        messages.error(request, 'You do not have permission to reserve books.')
        return redirect('book_detail', book_id=book.id)

    # Check if book is available
    if book.is_available():
        messages.error(request, 'This book is currently available. You can borrow it directly.')
        return redirect('book_detail', book_id=book.id)

    # Check if user already has a reservation for this book
    if Reservation.objects.filter(book=book, user=request.user, status='active').exists():
        messages.error(request, 'You already have an active reservation for this book.')
        return redirect('book_detail', book_id=book.id)

    if request.method == 'POST':
        form = ReservationForm(request.POST, user=request.user)
        if form.is_valid():
            reservation = form.save(commit=False)
            reservation.user = request.user
            reservation.book = book

            # Set priority (next in queue)
            last_reservation = Reservation.objects.filter(
                book=book,
                status='active'
            ).order_by('-priority').first()

            reservation.priority = (last_reservation.priority + 1) if last_reservation else 1
            reservation.save()

            # Log the reservation
            AuditLog.objects.create(
                action='reservation_made',
                user=request.user,
                book=book,
                description=f"Book '{book.title}' reserved by {request.user.get_full_name()}",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            messages.success(request, f'Book "{book.title}" reserved successfully! You are #{ reservation.priority} in the queue.')
            return redirect('book_detail', book_id=book.id)
    else:
        form = ReservationForm(user=request.user)
        form.fields['book'].initial = book
        form.fields['book'].widget.attrs['readonly'] = True

    return render(request, 'library/reservation_form.html', {'form': form, 'book': book})


@login_required
def reservation_cancel(request, reservation_id):
    """Cancel a reservation"""
    reservation = get_object_or_404(Reservation, id=reservation_id)

    # Check permissions
    if reservation.user != request.user and not request.user.can_manage_books():
        messages.error(request, 'You do not have permission to cancel this reservation.')
        return redirect('profile')

    if request.method == 'POST':
        if reservation.status != 'active':
            messages.error(request, 'This reservation is not active.')
            return redirect('profile')

        book = reservation.book
        reservation.status = 'cancelled'
        reservation.save()

        # Update priorities for remaining reservations
        remaining_reservations = Reservation.objects.filter(
            book=book,
            status='active',
            priority__gt=reservation.priority
        ).order_by('priority')

        for i, res in enumerate(remaining_reservations, start=reservation.priority):
            res.priority = i
            res.save()

        # Log the cancellation
        AuditLog.objects.create(
            action='reservation_cancelled',
            user=request.user,
            target_user=reservation.user,
            book=book,
            description=f"Reservation for '{book.title}' cancelled by {request.user.get_full_name()}",
            ip_address=request.META.get('REMOTE_ADDR')
        )

        messages.success(request, f'Reservation for "{book.title}" cancelled successfully!')

        if request.user.can_manage_books():
            return redirect('book_detail', book_id=book.id)
        else:
            return redirect('profile')

    return render(request, 'library/reservation_cancel_confirm.html', {'reservation': reservation})


@user_passes_test(is_librarian_or_admin)
def reports_dashboard(request):
    """Reports dashboard with various statistics"""
    from django.db.models import Count, Q
    from datetime import datetime, timedelta

    # Date range for reports (last 30 days by default)
    end_date = timezone.now()
    start_date = end_date - timedelta(days=30)

    # Basic statistics
    stats = {
        'total_books': Book.objects.count(),
        'total_users': User.objects.filter(is_active_member=True).count(),
        'active_loans': Loan.objects.filter(status='active').count(),
        'overdue_loans': Loan.objects.filter(status='overdue').count(),
        'total_reservations': Reservation.objects.filter(status='active').count(),
    }

    # Loan statistics for the period
    loans_in_period = Loan.objects.filter(issue_date__range=[start_date, end_date])
    stats.update({
        'loans_issued': loans_in_period.count(),
        'books_returned': loans_in_period.filter(status='returned').count(),
        'average_loan_duration': 14,  # Calculate actual average
    })

    # Popular books (most borrowed)
    popular_books = Book.objects.annotate(
        loan_count=Count('loans')
    ).order_by('-loan_count')[:10]

    # Active users (most loans)
    active_users = User.objects.annotate(
        loan_count=Count('loans', filter=Q(loans__issue_date__range=[start_date, end_date]))
    ).filter(loan_count__gt=0).order_by('-loan_count')[:10]

    # Category statistics
    category_stats = Category.objects.annotate(
        book_count=Count('books'),
        loan_count=Count('books__loans', filter=Q(books__loans__issue_date__range=[start_date, end_date]))
    ).order_by('-loan_count')

    # Monthly loan trends (last 12 months)
    monthly_loans = []
    for i in range(12):
        month_start = (end_date - timedelta(days=30*i)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        loan_count = Loan.objects.filter(issue_date__range=[month_start, month_end]).count()
        monthly_loans.append({
            'month': month_start.strftime('%b %Y'),
            'loans': loan_count
        })
    monthly_loans.reverse()

    context = {
        'stats': stats,
        'popular_books': popular_books,
        'active_users': active_users,
        'category_stats': category_stats,
        'monthly_loans': monthly_loans,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'library/reports_dashboard.html', context)


@user_passes_test(is_librarian_or_admin)
def export_loans_report(request):
    """Export loans report to Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from datetime import datetime

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loans Report"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = [
        'Loan ID', 'Book Title', 'ISBN', 'Borrower Name', 'Borrower Role',
        'Issue Date', 'Due Date', 'Return Date', 'Status', 'Days Overdue', 'Fine Amount'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Get loans data
    loans = Loan.objects.select_related('book', 'borrower').order_by('-issue_date')

    # Add data rows
    for row, loan in enumerate(loans, 2):
        ws.cell(row=row, column=1, value=str(loan.id))
        ws.cell(row=row, column=2, value=loan.book.title)
        ws.cell(row=row, column=3, value=loan.book.isbn)
        ws.cell(row=row, column=4, value=loan.borrower.get_full_name())
        ws.cell(row=row, column=5, value=loan.borrower.get_role_display())
        ws.cell(row=row, column=6, value=loan.issue_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=7, value=loan.due_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=8, value=loan.return_date.strftime('%Y-%m-%d') if loan.return_date else 'Not Returned')
        ws.cell(row=row, column=9, value=loan.get_status_display())
        ws.cell(row=row, column=10, value=loan.days_overdue() if loan.is_overdue() else 0)
        ws.cell(row=row, column=11, value=float(loan.fine_amount))

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="loans_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'

    wb.save(response)
    return response


@user_passes_test(is_librarian_or_admin)
def export_books_report(request):
    """Export books report to PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from django.http import HttpResponse
    from datetime import datetime

    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="books_report_{datetime.now().strftime("%Y%m%d")}.pdf"'

    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.5*inch)
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )

    # Title
    title = Paragraph("Library Books Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))

    # Report info
    report_info = Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal'])
    elements.append(report_info)
    elements.append(Spacer(1, 20))

    # Books data
    books = Book.objects.select_related('publisher').prefetch_related('authors', 'categories')

    # Table data
    data = [['Title', 'Authors', 'ISBN', 'Publisher', 'Status', 'Copies']]

    for book in books:
        data.append([
            book.title[:30] + '...' if len(book.title) > 30 else book.title,
            book.get_authors_display()[:25] + '...' if len(book.get_authors_display()) > 25 else book.get_authors_display(),
            book.isbn,
            str(book.publisher)[:20] + '...' if book.publisher and len(str(book.publisher)) > 20 else str(book.publisher or 'N/A'),
            book.get_status_display(),
            f"{book.available_copies}/{book.total_copies}"
        ])

    # Create table
    table = Table(data, colWidths=[2.5*inch, 1.5*inch, 1*inch, 1.2*inch, 0.8*inch, 0.7*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)

    # Build PDF
    doc.build(elements)
    return response


def advanced_search(request):
    """Advanced search functionality"""
    books = Book.objects.select_related('publisher').prefetch_related('authors', 'categories')
    users = User.objects.none()
    loans = Loan.objects.none()

    search_type = request.GET.get('type', 'books')
    query = request.GET.get('q', '')

    if query:
        if search_type == 'books':
            books = books.filter(
                Q(title__icontains=query) |
                Q(subtitle__icontains=query) |
                Q(authors__first_name__icontains=query) |
                Q(authors__last_name__icontains=query) |
                Q(isbn__icontains=query) |
                Q(publisher__name__icontains=query) |
                Q(categories__name__icontains=query) |
                Q(summary__icontains=query)
            ).distinct()

        elif search_type == 'users':
            if request.user.can_manage_users():
                users = User.objects.filter(
                    Q(username__icontains=query) |
                    Q(first_name__icontains=query) |
                    Q(last_name__icontains=query) |
                    Q(email__icontains=query) |
                    Q(enrollment_number__icontains=query)
                ).distinct()

        elif search_type == 'loans':
            if request.user.can_manage_books():
                loans = Loan.objects.select_related('book', 'borrower').filter(
                    Q(book__title__icontains=query) |
                    Q(borrower__first_name__icontains=query) |
                    Q(borrower__last_name__icontains=query) |
                    Q(borrower__username__icontains=query)
                ).distinct()

    # Pagination
    if search_type == 'books':
        paginator = Paginator(books, 12)
        items = books
    elif search_type == 'users':
        paginator = Paginator(users, 20)
        items = users
    elif search_type == 'loans':
        paginator = Paginator(loans, 20)
        items = loans
    else:
        paginator = Paginator([], 20)
        items = []

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'query': query,
        'search_type': search_type,
        'page_obj': page_obj,
        'items': page_obj,
        'total_results': paginator.count if hasattr(paginator, 'count') else len(items),
    }

    return render(request, 'library/advanced_search.html', context)


@user_passes_test(is_admin)
def user_management(request):
    """User management page with multiple import options"""
    return render(request, 'library/user_management.html')


@user_passes_test(is_admin)
def download_csv_template(request):
    """Download CSV template for user import"""
    import csv
    from django.http import HttpResponse

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="user_import_template.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'username', 'email', 'first_name', 'last_name', 'role',
        'phone_number', 'date_of_birth', 'address', 'enrollment_number', 'class_grade'
    ])

    # Add sample data
    writer.writerow([
        'john.doe', 'john.doe@school.edu', 'John', 'Doe', 'student',
        '+1-555-0123', '2005-01-15', '123 Main St, City, State', 'STU001', '10A'
    ])
    writer.writerow([
        'jane.smith', 'jane.smith@school.edu', 'Jane', 'Smith', 'teacher',
        '+1-555-0124', '1985-03-20', '456 Oak Ave, City, State', 'TCH001', ''
    ])

    return response


@user_passes_test(is_admin)
def import_users_csv(request):
    """Import users from CSV file"""
    if request.method == 'POST':
        csv_file = request.FILES.get('csv_file')
        send_emails = request.POST.get('send_welcome_emails') == 'on'

        if not csv_file:
            messages.error(request, 'Please select a CSV file.')
            return redirect('user_management')

        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'Please upload a valid CSV file.')
            return redirect('user_management')

        try:
            import csv
            import io
            from django.contrib.auth.hashers import make_password

            # Read CSV file
            file_data = csv_file.read().decode('utf-8')
            csv_data = csv.DictReader(io.StringIO(file_data))

            created_users = []
            errors = []

            for row_num, row in enumerate(csv_data, start=2):
                try:
                    # Validate required fields
                    required_fields = ['username', 'email', 'first_name', 'last_name', 'role']
                    for field in required_fields:
                        if not row.get(field, '').strip():
                            errors.append(f"Row {row_num}: Missing required field '{field}'")
                            continue

                    # Check if user already exists
                    if User.objects.filter(username=row['username']).exists():
                        errors.append(f"Row {row_num}: Username '{row['username']}' already exists")
                        continue

                    if User.objects.filter(email=row['email']).exists():
                        errors.append(f"Row {row_num}: Email '{row['email']}' already exists")
                        continue

                    # Create user
                    user_data = {
                        'username': row['username'].strip(),
                        'email': row['email'].strip(),
                        'first_name': row['first_name'].strip(),
                        'last_name': row['last_name'].strip(),
                        'role': row['role'].strip(),
                        'password': make_password('temp123'),  # Temporary password
                        'is_active_member': True,
                    }

                    # Add optional fields
                    optional_fields = ['phone_number', 'address', 'enrollment_number', 'class_grade']
                    for field in optional_fields:
                        if row.get(field, '').strip():
                            user_data[field] = row[field].strip()

                    # Handle date of birth
                    if row.get('date_of_birth', '').strip():
                        from datetime import datetime
                        try:
                            user_data['date_of_birth'] = datetime.strptime(row['date_of_birth'], '%Y-%m-%d').date()
                        except ValueError:
                            errors.append(f"Row {row_num}: Invalid date format for date_of_birth (use YYYY-MM-DD)")
                            continue

                    user = User.objects.create(**user_data)
                    created_users.append(user)

                    # Log user creation
                    AuditLog.objects.create(
                        action='user_created',
                        user=request.user,
                        target_user=user,
                        description=f"User {user.username} created via CSV import",
                        ip_address=request.META.get('REMOTE_ADDR')
                    )

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")

            # Send welcome emails if requested
            if send_emails and created_users:
                from django.core.mail import send_mail
                from django.conf import settings

                for user in created_users:
                    try:
                        send_mail(
                            'Welcome to LibraryPro',
                            f'Hello {user.get_full_name()},\n\nYour library account has been created.\n\nUsername: {user.username}\nTemporary Password: temp123\n\nPlease log in and change your password.\n\nBest regards,\nLibrary Team',
                            settings.DEFAULT_FROM_EMAIL,
                            [user.email],
                            fail_silently=True,
                        )
                    except:
                        pass  # Continue even if email fails

            # Show results
            if created_users:
                messages.success(request, f'Successfully imported {len(created_users)} users.')

            if errors:
                error_message = f'{len(errors)} errors occurred:\n' + '\n'.join(errors[:10])
                if len(errors) > 10:
                    error_message += f'\n... and {len(errors) - 10} more errors.'
                messages.error(request, error_message)

        except Exception as e:
            messages.error(request, f'Error processing CSV file: {str(e)}')

    return redirect('user_management')


@user_passes_test(is_admin)
def sync_school_api(request):
    """Sync users from school management system API"""
    if request.method == 'POST':
        api_url = request.POST.get('api_url')
        api_key = request.POST.get('api_key')

        if not api_url or not api_key:
            messages.error(request, 'Please provide both API URL and API key.')
            return redirect('user_management')

        try:
            import requests
            from django.contrib.auth.hashers import make_password

            # Make API request
            headers = {'Authorization': f'Bearer {api_key}'}
            response = requests.get(api_url, headers=headers, timeout=30)
            response.raise_for_status()

            data = response.json()
            users_data = data.get('users', [])

            if not users_data:
                messages.warning(request, 'No users found in API response.')
                return redirect('user_management')

            created_users = []
            updated_users = []
            errors = []

            for user_data in users_data:
                try:
                    username = user_data.get('username')
                    email = user_data.get('email')

                    if not username or not email:
                        errors.append(f"Missing username or email for user: {user_data}")
                        continue

                    # Check if user exists
                    user, created = User.objects.get_or_create(
                        username=username,
                        defaults={
                            'email': email,
                            'first_name': user_data.get('first_name', ''),
                            'last_name': user_data.get('last_name', ''),
                            'role': user_data.get('role', 'student'),
                            'password': make_password('temp123'),
                            'is_active_member': True,
                            'phone_number': user_data.get('phone_number', ''),
                            'enrollment_number': user_data.get('enrollment_number', ''),
                            'class_grade': user_data.get('class_grade', ''),
                        }
                    )

                    if created:
                        created_users.append(user)
                        # Log user creation
                        AuditLog.objects.create(
                            action='user_created',
                            user=request.user,
                            target_user=user,
                            description=f"User {user.username} created via API sync",
                            ip_address=request.META.get('REMOTE_ADDR')
                        )
                    else:
                        # Update existing user
                        user.email = email
                        user.first_name = user_data.get('first_name', user.first_name)
                        user.last_name = user_data.get('last_name', user.last_name)
                        user.role = user_data.get('role', user.role)
                        user.phone_number = user_data.get('phone_number', user.phone_number)
                        user.enrollment_number = user_data.get('enrollment_number', user.enrollment_number)
                        user.class_grade = user_data.get('class_grade', user.class_grade)
                        user.save()
                        updated_users.append(user)

                except Exception as e:
                    errors.append(f"Error processing user {user_data.get('username', 'unknown')}: {str(e)}")

            # Show results
            result_messages = []
            if created_users:
                result_messages.append(f'Created {len(created_users)} new users')
            if updated_users:
                result_messages.append(f'Updated {len(updated_users)} existing users')

            if result_messages:
                messages.success(request, '. '.join(result_messages) + '.')

            if errors:
                error_message = f'{len(errors)} errors occurred during sync.'
                messages.error(request, error_message)

        except requests.RequestException as e:
            messages.error(request, f'API connection error: {str(e)}')
        except Exception as e:
            messages.error(request, f'Error syncing with API: {str(e)}')

    return redirect('user_management')


def book_list(request):
    """List all books with search and filtering"""
    form = BookSearchForm(request.GET)
    books = Book.objects.select_related('publisher').prefetch_related('authors', 'categories')

    if form.is_valid():
        query = form.cleaned_data.get('query')
        category = form.cleaned_data.get('category')
        status = form.cleaned_data.get('status')
        section = form.cleaned_data.get('section')

        if query:
            books = books.filter(
                Q(title__icontains=query) |
                Q(authors__first_name__icontains=query) |
                Q(authors__last_name__icontains=query) |
                Q(isbn__icontains=query) |
                Q(publisher__name__icontains=query)
            ).distinct()

        if category:
            books = books.filter(categories=category)

        if status:
            books = books.filter(status=status)

        if section:
            books = books.filter(section__icontains=section)

    # Pagination
    paginator = Paginator(books, 12)  # 12 books per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'form': form,
        'page_obj': page_obj,
        'books': page_obj,
    }

    return render(request, 'library/book_list.html', context)


def book_detail(request, book_id):
    """Book detail view"""
    book = get_object_or_404(Book, id=book_id)

    # Check if user can reserve this book
    can_reserve = False
    if request.user.is_authenticated and request.user.can_borrow_books():
        can_reserve = (
            not book.is_available() and
            not Reservation.objects.filter(
                book=book,
                user=request.user,
                status='active'
            ).exists()
        )

    # Get active reservations for this book
    reservations = Reservation.objects.filter(
        book=book,
        status='active'
    ).select_related('user').order_by('priority')

    # Get loan history (for librarians/admins)
    loan_history = None
    if request.user.is_authenticated and request.user.can_manage_books():
        loan_history = Loan.objects.filter(book=book).select_related('borrower').order_by('-issue_date')[:10]

    context = {
        'book': book,
        'can_reserve': can_reserve,
        'reservations': reservations,
        'loan_history': loan_history,
    }

    return render(request, 'library/book_detail.html', context)


@user_passes_test(can_manage_books)
def book_add(request):
    """Add new book"""
    if request.method == 'POST':
        form = BookForm(request.POST, request.FILES)
        if form.is_valid():
            book = form.save(commit=False)
            book.created_by = request.user
            book.save()
            form.save_m2m()

            # Log the book addition
            AuditLog.objects.create(
                action='book_added',
                user=request.user,
                book=book,
                description=f"Book '{book.title}' added by {request.user.username}",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            messages.success(request, f'Book "{book.title}" added successfully!')
            return redirect('book_detail', book_id=book.id)
    else:
        form = BookForm()

    return render(request, 'library/book_form.html', {'form': form, 'title': 'Add Book'})


@user_passes_test(can_manage_books)
def book_edit(request, book_id):
    """Edit existing book"""
    book = get_object_or_404(Book, id=book_id)

    if request.method == 'POST':
        form = BookForm(request.POST, request.FILES, instance=book)
        if form.is_valid():
            book = form.save()

            # Log the book update
            AuditLog.objects.create(
                action='book_updated',
                user=request.user,
                book=book,
                description=f"Book '{book.title}' updated by {request.user.username}",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            messages.success(request, f'Book "{book.title}" updated successfully!')
            return redirect('book_detail', book_id=book.id)
    else:
        form = BookForm(instance=book)

    return render(request, 'library/book_form.html', {'form': form, 'title': 'Edit Book', 'book': book})


@user_passes_test(can_manage_books)
def book_delete(request, book_id):
    """Delete book"""
    book = get_object_or_404(Book, id=book_id)

    if request.method == 'POST':
        # Check if book has active loans
        if Loan.objects.filter(book=book, status='active').exists():
            messages.error(request, 'Cannot delete book with active loans.')
            return redirect('book_detail', book_id=book.id)

        book_title = book.title

        # Log the book deletion
        AuditLog.objects.create(
            action='book_deleted',
            user=request.user,
            description=f"Book '{book_title}' deleted by {request.user.username}",
            ip_address=request.META.get('REMOTE_ADDR')
        )

        book.delete()
        messages.success(request, f'Book "{book_title}" deleted successfully!')
        return redirect('book_list')

    return render(request, 'library/book_confirm_delete.html', {'book': book})


@user_passes_test(is_admin)
def library_settings(request):
    """Library settings management"""
    settings, created = LibrarySettings.objects.get_or_create(pk=1)

    if request.method == 'POST':
        form = LibrarySettingsForm(request.POST, request.FILES, instance=settings)
        if form.is_valid():
            form.save()

            # Log settings update
            AuditLog.objects.create(
                action='settings_updated',
                user=request.user,
                description=f"Library settings updated by {request.user.username}",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            messages.success(request, 'Library settings updated successfully!')
            return redirect('library_settings')
    else:
        form = LibrarySettingsForm(instance=settings)

    return render(request, 'library/library_settings.html', {'form': form, 'settings': settings})


# API Endpoints for Autocomplete
@login_required
def publisher_autocomplete(request):
    """Autocomplete for publishers"""
    if request.method == 'POST':
        # Create new publisher
        try:
            data = json.loads(request.body)
            publisher = Publisher.objects.create(
                name=data['name'],
                address=data.get('address', ''),
                email=data.get('email', ''),
                phone=data.get('phone', ''),
                website=data.get('website', '')
            )
            return JsonResponse({
                'id': publisher.id,
                'name': publisher.name,
                'address': publisher.address or '',
                'email': publisher.email or '',
                'phone': publisher.phone or '',
                'website': publisher.website or ''
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    # GET request - autocomplete
    query = request.GET.get('q', '')
    publishers = Publisher.objects.filter(name__icontains=query)[:10]

    results = []
    for publisher in publishers:
        results.append({
            'id': publisher.id,
            'name': publisher.name,
            'address': publisher.address or '',
            'email': publisher.email or '',
            'phone': publisher.phone or '',
            'website': publisher.website or ''
        })

    return JsonResponse({'results': results})


@login_required
def category_autocomplete(request):
    """Autocomplete for categories"""
    if request.method == 'POST':
        # Create new category
        try:
            data = json.loads(request.body)
            category = Category.objects.create(
                name=data['name'],
                description=data.get('description', '')
            )
            return JsonResponse({
                'id': category.id,
                'name': category.name,
                'description': category.description or ''
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    # GET request - autocomplete
    query = request.GET.get('q', '')
    categories = Category.objects.filter(name__icontains=query)[:10]

    results = []
    for category in categories:
        results.append({
            'id': category.id,
            'name': category.name,
            'description': category.description or ''
        })

    return JsonResponse({'results': results})


@login_required
def section_autocomplete(request):
    """Autocomplete for sections"""
    if request.method == 'POST':
        # Create new section
        try:
            data = json.loads(request.body)
            section = Section.objects.create(
                name=data['name'],
                description=data.get('description', ''),
                floor=data.get('floor', 'Ground Floor')
            )
            return JsonResponse({
                'id': section.id,
                'name': section.name,
                'description': section.description or '',
                'floor': section.floor
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    # GET request - autocomplete
    query = request.GET.get('q', '')
    sections = Section.objects.filter(name__icontains=query)[:10]

    results = []
    for section in sections:
        results.append({
            'id': section.id,
            'name': section.name,
            'description': section.description or '',
            'floor': section.floor
        })

    return JsonResponse({'results': results})


@login_required
def shelf_location_autocomplete(request):
    """Autocomplete for shelf locations"""
    query = request.GET.get('q', '')
    section_id = request.GET.get('section_id', '')

    shelf_locations = ShelfLocation.objects.filter(code__icontains=query)
    if section_id:
        shelf_locations = shelf_locations.filter(section_id=section_id)

    shelf_locations = shelf_locations[:10]

    results = []
    for shelf in shelf_locations:
        results.append({
            'id': shelf.id,
            'code': shelf.code,
            'section_name': shelf.section.name,
            'description': shelf.description or '',
            'capacity': shelf.capacity
        })

    return JsonResponse({'results': results})


@login_required
def floor_autocomplete(request):
    """Autocomplete for floors"""
    query = request.GET.get('q', '')
    floors = Floor.objects.filter(name__icontains=query)[:10]

    results = []
    for floor in floors:
        results.append({
            'id': floor.id,
            'name': floor.name,
            'description': floor.description or ''
        })

    return JsonResponse({'results': results})


# Management Views for Authors, Categories, Publishers, etc.

@user_passes_test(can_manage_books)
def author_list(request):
    """List all authors"""
    authors = Author.objects.all().order_by('last_name', 'first_name')

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        authors = authors.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(biography__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(authors, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Prepare data for template
    items = []
    for author in page_obj:
        items.append({
            'display_fields': [
                f"{author.first_name} {author.last_name}",
                author.nationality or 'N/A',
                author.birth_date.strftime('%Y-%m-%d') if author.birth_date else 'N/A',
                author.books.count()
            ],
            'edit_url': reverse('author_edit', args=[author.id]),
            'delete_url': reverse('author_delete', args=[author.id])
        })

    context = {
        'title': 'Authors',
        'description': 'Manage book authors',
        'singular_name': 'Author',
        'items': items,
        'page_obj': page_obj,
        'table_headers': ['Name', 'Nationality', 'Birth Date', 'Books Count'],
        'add_url': reverse('author_add'),
        'list_url': reverse('author_list'),
        'empty_icon': 'fa-user-edit',
        'empty_message': 'Start by adding your first author to organize your book collection.'
    }

    return render(request, 'library/management_list.html', context)


@user_passes_test(can_manage_books)
def author_add(request):
    """Add new author"""
    from django import forms

    class AuthorForm(forms.ModelForm):
        class Meta:
            model = Author
            fields = ['first_name', 'last_name', 'nationality', 'birth_date', 'death_date', 'biography']
            widgets = {
                'birth_date': forms.DateInput(attrs={'type': 'date'}),
                'death_date': forms.DateInput(attrs={'type': 'date'}),
                'biography': forms.Textarea(attrs={'rows': 4})
            }

    if request.method == 'POST':
        form = AuthorForm(request.POST)
        if form.is_valid():
            author = form.save()
            messages.success(request, f'Author "{author.first_name} {author.last_name}" added successfully!')
            return redirect('author_list')
    else:
        form = AuthorForm()

    context = {
        'form': form,
        'title': 'Add Author',
        'description': 'Add a new author to the library system',
        'back_url': reverse('author_list'),
        'submit_text': 'Add Author'
    }

    return render(request, 'library/management_form.html', context)


@user_passes_test(can_manage_books)
def author_edit(request, author_id):
    """Edit existing author"""
    from django import forms

    author = get_object_or_404(Author, id=author_id)

    class AuthorForm(forms.ModelForm):
        class Meta:
            model = Author
            fields = ['first_name', 'last_name', 'nationality', 'birth_date', 'death_date', 'biography']
            widgets = {
                'birth_date': forms.DateInput(attrs={'type': 'date'}),
                'death_date': forms.DateInput(attrs={'type': 'date'}),
                'biography': forms.Textarea(attrs={'rows': 4})
            }

    if request.method == 'POST':
        form = AuthorForm(request.POST, instance=author)
        if form.is_valid():
            author = form.save()
            messages.success(request, f'Author "{author.first_name} {author.last_name}" updated successfully!')
            return redirect('author_list')
    else:
        form = AuthorForm(instance=author)

    context = {
        'form': form,
        'title': 'Edit Author',
        'description': f'Edit details for {author.first_name} {author.last_name}',
        'back_url': reverse('author_list'),
        'submit_text': 'Update Author'
    }

    return render(request, 'library/management_form.html', context)


@user_passes_test(can_manage_books)
def author_delete(request, author_id):
    """Delete author"""
    author = get_object_or_404(Author, id=author_id)

    if request.method == 'POST':
        # Check if author has books
        if author.books.exists():
            messages.error(request, f'Cannot delete author "{author.first_name} {author.last_name}" because they have books associated with them.')
            return redirect('author_list')

        author_name = f"{author.first_name} {author.last_name}"
        author.delete()
        messages.success(request, f'Author "{author_name}" deleted successfully!')
        return redirect('author_list')

    return render(request, 'library/confirm_delete.html', {
        'object': author,
        'object_name': f"{author.first_name} {author.last_name}",
        'back_url': reverse('author_list')
    })


# Category Management Views
@user_passes_test(can_manage_books)
def category_list(request):
    """List all categories"""
    categories = Category.objects.all().order_by('name')

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        categories = categories.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(categories, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Prepare data for template
    items = []
    for category in page_obj:
        items.append({
            'display_fields': [
                category.name,
                category.description or 'N/A',
                category.books.count()
            ],
            'edit_url': reverse('category_edit', args=[category.id]),
            'delete_url': reverse('category_delete', args=[category.id])
        })

    context = {
        'title': 'Categories',
        'description': 'Manage book categories',
        'singular_name': 'Category',
        'items': items,
        'page_obj': page_obj,
        'table_headers': ['Name', 'Description', 'Books Count'],
        'add_url': reverse('category_add'),
        'list_url': reverse('category_list'),
        'empty_icon': 'fa-tags',
        'empty_message': 'Start by adding your first category to organize your book collection.'
    }

    return render(request, 'library/management_list.html', context)


@user_passes_test(can_manage_books)
def category_add(request):
    """Add new category"""
    from django import forms

    class CategoryForm(forms.ModelForm):
        class Meta:
            model = Category
            fields = ['name', 'description']
            widgets = {
                'description': forms.Textarea(attrs={'rows': 3})
            }

    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Category "{category.name}" added successfully!')
            return redirect('category_list')
    else:
        form = CategoryForm()

    context = {
        'form': form,
        'title': 'Add Category',
        'description': 'Add a new book category',
        'back_url': reverse('category_list'),
        'submit_text': 'Add Category'
    }

    return render(request, 'library/management_form.html', context)


@user_passes_test(can_manage_books)
def category_edit(request, category_id):
    """Edit existing category"""
    from django import forms

    category = get_object_or_404(Category, id=category_id)

    class CategoryForm(forms.ModelForm):
        class Meta:
            model = Category
            fields = ['name', 'description']
            widgets = {
                'description': forms.Textarea(attrs={'rows': 3})
            }

    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Category "{category.name}" updated successfully!')
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)

    context = {
        'form': form,
        'title': 'Edit Category',
        'description': f'Edit details for {category.name}',
        'back_url': reverse('category_list'),
        'submit_text': 'Update Category'
    }

    return render(request, 'library/management_form.html', context)


@user_passes_test(can_manage_books)
def category_delete(request, category_id):
    """Delete category"""
    category = get_object_or_404(Category, id=category_id)

    if request.method == 'POST':
        # Check if category has books
        if category.books.exists():
            messages.error(request, f'Cannot delete category "{category.name}" because it has books associated with it.')
            return redirect('category_list')

        category_name = category.name
        category.delete()
        messages.success(request, f'Category "{category_name}" deleted successfully!')
        return redirect('category_list')

    return render(request, 'library/confirm_delete.html', {
        'object': category,
        'object_name': category.name,
        'back_url': reverse('category_list')
    })


# Publisher Management Views
@user_passes_test(can_manage_books)
def publisher_list(request):
    """List all publishers"""
    publishers = Publisher.objects.all().order_by('name')

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        publishers = publishers.filter(
            Q(name__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(website__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(publishers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Prepare data for template
    items = []
    for publisher in page_obj:
        items.append({
            'display_fields': [
                publisher.name,
                publisher.address or 'N/A',
                publisher.website or 'N/A',
                publisher.books.count()
            ],
            'edit_url': reverse('publisher_edit', args=[publisher.id]),
            'delete_url': reverse('publisher_delete', args=[publisher.id])
        })

    context = {
        'title': 'Publishers',
        'description': 'Manage book publishers',
        'singular_name': 'Publisher',
        'items': items,
        'page_obj': page_obj,
        'table_headers': ['Name', 'Address', 'Website', 'Books Count'],
        'add_url': reverse('publisher_add'),
        'list_url': reverse('publisher_list'),
        'empty_icon': 'fa-building',
        'empty_message': 'Start by adding your first publisher to organize your book collection.'
    }

    return render(request, 'library/management_list.html', context)


@user_passes_test(can_manage_books)
def publisher_add(request):
    """Add new publisher"""
    from django import forms

    class PublisherForm(forms.ModelForm):
        class Meta:
            model = Publisher
            fields = ['name', 'address', 'phone', 'email', 'website']
            widgets = {
                'address': forms.Textarea(attrs={'rows': 3})
            }

    if request.method == 'POST':
        form = PublisherForm(request.POST)
        if form.is_valid():
            publisher = form.save()
            messages.success(request, f'Publisher "{publisher.name}" added successfully!')
            return redirect('publisher_list')
    else:
        form = PublisherForm()

    context = {
        'form': form,
        'title': 'Add Publisher',
        'description': 'Add a new book publisher',
        'back_url': reverse('publisher_list'),
        'submit_text': 'Add Publisher'
    }

    return render(request, 'library/management_form.html', context)


@user_passes_test(can_manage_books)
def publisher_edit(request, publisher_id):
    """Edit existing publisher"""
    from django import forms

    publisher = get_object_or_404(Publisher, id=publisher_id)

    class PublisherForm(forms.ModelForm):
        class Meta:
            model = Publisher
            fields = ['name', 'address', 'phone', 'email', 'website']
            widgets = {
                'address': forms.Textarea(attrs={'rows': 3})
            }

    if request.method == 'POST':
        form = PublisherForm(request.POST, instance=publisher)
        if form.is_valid():
            publisher = form.save()
            messages.success(request, f'Publisher "{publisher.name}" updated successfully!')
            return redirect('publisher_list')
    else:
        form = PublisherForm(instance=publisher)

    context = {
        'form': form,
        'title': 'Edit Publisher',
        'description': f'Edit details for {publisher.name}',
        'back_url': reverse('publisher_list'),
        'submit_text': 'Update Publisher'
    }

    return render(request, 'library/management_form.html', context)


@user_passes_test(can_manage_books)
def publisher_delete(request, publisher_id):
    """Delete publisher"""
    publisher = get_object_or_404(Publisher, id=publisher_id)

    if request.method == 'POST':
        # Check if publisher has books
        if publisher.books.exists():
            messages.error(request, f'Cannot delete publisher "{publisher.name}" because it has books associated with it.')
            return redirect('publisher_list')

        publisher_name = publisher.name
        publisher.delete()
        messages.success(request, f'Publisher "{publisher_name}" deleted successfully!')
        return redirect('publisher_list')

    return render(request, 'library/confirm_delete.html', {
        'object': publisher,
        'object_name': publisher.name,
        'back_url': reverse('publisher_list')
    })


# Section Management Views
@user_passes_test(can_manage_books)
def section_list(request):
    """List all sections"""
    sections = Section.objects.all().order_by('name')

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        sections = sections.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(sections, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Prepare data for template
    items = []
    for section in page_obj:
        items.append({
            'display_fields': [
                section.name,
                section.description or 'N/A',
                section.books.count()
            ],
            'edit_url': reverse('section_edit', args=[section.id]),
            'delete_url': reverse('section_delete', args=[section.id])
        })

    context = {
        'title': 'Sections',
        'description': 'Manage library sections',
        'singular_name': 'Section',
        'items': items,
        'page_obj': page_obj,
        'table_headers': ['Name', 'Description', 'Books Count'],
        'add_url': reverse('section_add'),
        'list_url': reverse('section_list'),
        'empty_icon': 'fa-layer-group',
        'empty_message': 'Start by adding your first section to organize your library.'
    }

    return render(request, 'library/management_list.html', context)


@user_passes_test(can_manage_books)
def section_add(request):
    """Add new section"""
    from django import forms

    class SectionForm(forms.ModelForm):
        class Meta:
            model = Section
            fields = ['name', 'description']
            widgets = {
                'description': forms.Textarea(attrs={'rows': 3})
            }

    if request.method == 'POST':
        form = SectionForm(request.POST)
        if form.is_valid():
            section = form.save()
            messages.success(request, f'Section "{section.name}" added successfully!')
            return redirect('section_list')
    else:
        form = SectionForm()

    context = {
        'form': form,
        'title': 'Add Section',
        'description': 'Add a new library section',
        'back_url': reverse('section_list'),
        'submit_text': 'Add Section'
    }

    return render(request, 'library/management_form.html', context)


@user_passes_test(can_manage_books)
def section_edit(request, section_id):
    """Edit existing section"""
    from django import forms

    section = get_object_or_404(Section, id=section_id)

    class SectionForm(forms.ModelForm):
        class Meta:
            model = Section
            fields = ['name', 'description']
            widgets = {
                'description': forms.Textarea(attrs={'rows': 3})
            }

    if request.method == 'POST':
        form = SectionForm(request.POST, instance=section)
        if form.is_valid():
            section = form.save()
            messages.success(request, f'Section "{section.name}" updated successfully!')
            return redirect('section_list')
    else:
        form = SectionForm(instance=section)

    context = {
        'form': form,
        'title': 'Edit Section',
        'description': f'Edit details for {section.name}',
        'back_url': reverse('section_list'),
        'submit_text': 'Update Section'
    }

    return render(request, 'library/management_form.html', context)


@user_passes_test(can_manage_books)
def section_delete(request, section_id):
    """Delete section"""
    section = get_object_or_404(Section, id=section_id)

    if request.method == 'POST':
        # Check if section has books
        if section.books.exists():
            messages.error(request, f'Cannot delete section "{section.name}" because it has books associated with it.')
            return redirect('section_list')

        section_name = section.name
        section.delete()
        messages.success(request, f'Section "{section_name}" deleted successfully!')
        return redirect('section_list')

    return render(request, 'library/confirm_delete.html', {
        'object': section,
        'object_name': section.name,
        'back_url': reverse('section_list')
    })


# Shelf Location Management Views
@user_passes_test(can_manage_books)
def shelf_location_list(request):
    """List all shelf locations"""
    shelves = ShelfLocation.objects.all().order_by('code')

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        shelves = shelves.filter(
            Q(code__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(shelves, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Prepare data for template
    items = []
    for shelf in page_obj:
        items.append({
            'display_fields': [
                shelf.code,
                shelf.description or 'N/A',
                shelf.books.count()
            ],
            'edit_url': reverse('shelf_location_edit', args=[shelf.id]),
            'delete_url': reverse('shelf_location_delete', args=[shelf.id])
        })

    context = {
        'title': 'Shelf Locations',
        'description': 'Manage library shelf locations',
        'singular_name': 'Shelf Location',
        'items': items,
        'page_obj': page_obj,
        'table_headers': ['Code', 'Description', 'Books Count'],
        'add_url': reverse('shelf_location_add'),
        'list_url': reverse('shelf_location_list'),
        'empty_icon': 'fa-map-marker-alt',
        'empty_message': 'Start by adding your first shelf location to organize your library.'
    }

    return render(request, 'library/management_list.html', context)


@user_passes_test(can_manage_books)
def shelf_location_add(request):
    """Add new shelf location"""
    from django import forms

    class ShelfLocationForm(forms.ModelForm):
        class Meta:
            model = ShelfLocation
            fields = ['code', 'description']
            widgets = {
                'description': forms.Textarea(attrs={'rows': 3})
            }

    if request.method == 'POST':
        form = ShelfLocationForm(request.POST)
        if form.is_valid():
            shelf = form.save()
            messages.success(request, f'Shelf location "{shelf.code}" added successfully!')
            return redirect('shelf_location_list')
    else:
        form = ShelfLocationForm()

    context = {
        'form': form,
        'title': 'Add Shelf Location',
        'description': 'Add a new shelf location',
        'back_url': reverse('shelf_location_list'),
        'submit_text': 'Add Shelf Location'
    }

    return render(request, 'library/management_form.html', context)


@user_passes_test(can_manage_books)
def shelf_location_edit(request, shelf_id):
    """Edit existing shelf location"""
    from django import forms

    shelf = get_object_or_404(ShelfLocation, id=shelf_id)

    class ShelfLocationForm(forms.ModelForm):
        class Meta:
            model = ShelfLocation
            fields = ['code', 'description']
            widgets = {
                'description': forms.Textarea(attrs={'rows': 3})
            }

    if request.method == 'POST':
        form = ShelfLocationForm(request.POST, instance=shelf)
        if form.is_valid():
            shelf = form.save()
            messages.success(request, f'Shelf location "{shelf.code}" updated successfully!')
            return redirect('shelf_location_list')
    else:
        form = ShelfLocationForm(instance=shelf)

    context = {
        'form': form,
        'title': 'Edit Shelf Location',
        'description': f'Edit details for {shelf.code}',
        'back_url': reverse('shelf_location_list'),
        'submit_text': 'Update Shelf Location'
    }

    return render(request, 'library/management_form.html', context)


@user_passes_test(can_manage_books)
def shelf_location_delete(request, shelf_id):
    """Delete shelf location"""
    shelf = get_object_or_404(ShelfLocation, id=shelf_id)

    if request.method == 'POST':
        # Check if shelf has books
        if shelf.books.exists():
            messages.error(request, f'Cannot delete shelf location "{shelf.code}" because it has books associated with it.')
            return redirect('shelf_location_list')

        shelf_code = shelf.code
        shelf.delete()
        messages.success(request, f'Shelf location "{shelf_code}" deleted successfully!')
        return redirect('shelf_location_list')

    return render(request, 'library/confirm_delete.html', {
        'object': shelf,
        'object_name': shelf.code,
        'back_url': reverse('shelf_location_list')
    })


# Floor Management Views
@user_passes_test(can_manage_books)
def floor_list(request):
    """List all floors"""
    floors = Floor.objects.all().order_by('name')

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        floors = floors.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(floors, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Prepare data for template
    items = []
    for floor in page_obj:
        items.append({
            'display_fields': [
                floor.name,
                floor.description or 'N/A',
                floor.books.count()
            ],
            'edit_url': reverse('floor_edit', args=[floor.id]),
            'delete_url': reverse('floor_delete', args=[floor.id])
        })

    context = {
        'title': 'Floors',
        'description': 'Manage library floors',
        'singular_name': 'Floor',
        'items': items,
        'page_obj': page_obj,
        'table_headers': ['Name', 'Description', 'Books Count'],
        'add_url': reverse('floor_add'),
        'list_url': reverse('floor_list'),
        'empty_icon': 'fa-building',
        'empty_message': 'Start by adding your first floor to organize your library.'
    }

    return render(request, 'library/management_list.html', context)


@user_passes_test(can_manage_books)
def floor_add(request):
    """Add new floor"""
    from django import forms

    class FloorForm(forms.ModelForm):
        class Meta:
            model = Floor
            fields = ['name', 'description']
            widgets = {
                'description': forms.Textarea(attrs={'rows': 3})
            }

    if request.method == 'POST':
        form = FloorForm(request.POST)
        if form.is_valid():
            floor = form.save()
            messages.success(request, f'Floor "{floor.name}" added successfully!')
            return redirect('floor_list')
    else:
        form = FloorForm()

    context = {
        'form': form,
        'title': 'Add Floor',
        'description': 'Add a new library floor',
        'back_url': reverse('floor_list'),
        'submit_text': 'Add Floor'
    }

    return render(request, 'library/management_form.html', context)


@user_passes_test(can_manage_books)
def floor_edit(request, floor_id):
    """Edit existing floor"""
    from django import forms

    floor = get_object_or_404(Floor, id=floor_id)

    class FloorForm(forms.ModelForm):
        class Meta:
            model = Floor
            fields = ['name', 'description']
            widgets = {
                'description': forms.Textarea(attrs={'rows': 3})
            }

    if request.method == 'POST':
        form = FloorForm(request.POST, instance=floor)
        if form.is_valid():
            floor = form.save()
            messages.success(request, f'Floor "{floor.name}" updated successfully!')
            return redirect('floor_list')
    else:
        form = FloorForm(instance=floor)

    context = {
        'form': form,
        'title': 'Edit Floor',
        'description': f'Edit details for {floor.name}',
        'back_url': reverse('floor_list'),
        'submit_text': 'Update Floor'
    }

    return render(request, 'library/management_form.html', context)


@user_passes_test(can_manage_books)
def floor_delete(request, floor_id):
    """Delete floor"""
    floor = get_object_or_404(Floor, id=floor_id)

    if request.method == 'POST':
        # Check if floor has books
        if floor.books.exists():
            messages.error(request, f'Cannot delete floor "{floor.name}" because it has books associated with it.')
            return redirect('floor_list')

        floor_name = floor.name
        floor.delete()
        messages.success(request, f'Floor "{floor_name}" deleted successfully!')
        return redirect('floor_list')

    return render(request, 'library/confirm_delete.html', {
        'object': floor,
        'object_name': floor.name,
        'back_url': reverse('floor_list')
    })


@user_passes_test(can_manage_books)
def book_autocomplete(request):
    """Autocomplete for books"""
    query = request.GET.get('q', '')
    books = Book.objects.filter(
        Q(title__icontains=query) | Q(isbn__icontains=query),
        status='available',
        available_copies__gt=0
    ).select_related('publisher').prefetch_related('authors')[:10]

    data = []
    for book in books:
        authors = ', '.join([f"{author.first_name} {author.last_name}" for author in book.authors.all()])
        data.append({
            'id': str(book.id),
            'text': f"{book.title} - {authors}",
            'title': book.title,
            'authors': authors,
            'isbn': book.isbn,
            'available_copies': book.available_copies
        })

    return JsonResponse(data, safe=False)


@user_passes_test(can_manage_books)
def user_autocomplete(request):
    """Autocomplete for users who can borrow books"""
    query = request.GET.get('q', '')
    users = User.objects.filter(
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(username__icontains=query) |
        Q(enrollment_number__icontains=query),
        role__in=['student', 'teacher'],
        is_active_member=True
    )[:10]

    data = []
    for user in users:
        data.append({
            'id': str(user.id),
            'text': f"{user.get_full_name()} ({user.username})",
            'name': user.get_full_name(),
            'username': user.username,
            'enrollment_number': user.enrollment_number,
            'role': user.get_role_display()
        })

    return JsonResponse(data, safe=False)


@user_passes_test(is_librarian_or_admin)
def pending_requests(request):
    """View and manage pending loan requests"""
    pending_loans = Loan.objects.filter(
        status='pending'
    ).select_related('book', 'borrower').order_by('-issue_date')

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        pending_loans = pending_loans.filter(
            Q(book__title__icontains=search_query) |
            Q(borrower__username__icontains=search_query) |
            Q(borrower__first_name__icontains=search_query) |
            Q(borrower__last_name__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(pending_loans, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'pending_loans': page_obj,
        'search_query': search_query,
    }

    return render(request, 'library/pending_requests.html', context)


@user_passes_test(is_librarian_or_admin)
def approve_request(request, loan_id):
    """Approve a pending loan request"""
    loan = get_object_or_404(Loan, id=loan_id, status='pending')

    if request.method == 'POST':
        # Check if book is still available
        book = loan.book
        if not book.is_available():
            messages.error(request, 'This book is no longer available.')
            return redirect('pending_requests')

        # Approve the loan
        loan.status = 'active'
        loan.issued_by = request.user
        loan.issue_date = timezone.now()
        loan.save()

        # Update book availability
        book.available_copies -= 1
        book.update_availability()

        # Log the approval
        AuditLog.objects.create(
            action='loan_approved',
            user=request.user,
            target_user=loan.borrower,
            book=book,
            description=f"Loan request for '{book.title}' approved by {request.user.username}",
            ip_address=request.META.get('REMOTE_ADDR')
        )

        # Create notification for the borrower
        create_notification(
            user=loan.borrower,
            title="Book Request Approved",
            message=f"Your request for '{book.title}' has been approved. Please visit the library to collect your book.",
            notification_type='request_approved',
            book=book,
            loan=loan
        )

        # Send email notification to borrower and additional email
        send_email_notification(
            user=loan.borrower,
            subject="Book Request Approved",
            message=f"""Dear {loan.borrower.get_full_name()},

Your request for the book "{book.title}" has been approved.

Book Details:
- Title: {book.title}
- Due Date: {loan.due_date.strftime('%B %d, %Y')}

Please visit the library to collect your book.

Best regards,
Library Team"""
        )

        messages.success(request, f'Loan request approved. Book "{book.title}" issued to {loan.borrower.get_full_name()}.')
        return redirect('pending_requests')

    context = {
        'loan': loan,
    }

    return render(request, 'library/approve_request.html', context)


@user_passes_test(is_librarian_or_admin)
def reject_request(request, loan_id):
    """Reject a pending loan request"""
    loan = get_object_or_404(Loan, id=loan_id, status='pending')

    if request.method == 'POST':
        reason = request.POST.get('reason', '')

        # Delete the pending request
        book_title = loan.book.title
        borrower_name = loan.borrower.get_full_name()

        # Log the rejection
        AuditLog.objects.create(
            action='loan_rejected',
            user=request.user,
            target_user=loan.borrower,
            book=loan.book,
            description=f"Loan request for '{book_title}' rejected by {request.user.username}. Reason: {reason}",
            ip_address=request.META.get('REMOTE_ADDR')
        )

        # Create notification for the borrower
        create_notification(
            user=loan.borrower,
            title="Book Request Rejected",
            message=f"Your request for '{book_title}' has been rejected. Reason: {reason}",
            notification_type='request_rejected',
            book=loan.book,
            loan=loan
        )

        loan.delete()

        messages.success(request, f'Loan request for "{book_title}" by {borrower_name} has been rejected.')
        return redirect('pending_requests')

    context = {
        'loan': loan,
    }

    return render(request, 'library/reject_request.html', context)


@login_required
def cancel_request(request, loan_id):
    """Allow students to cancel their pending requests"""
    loan = get_object_or_404(Loan, id=loan_id, status='pending', borrower=request.user)

    if request.method == 'POST':
        book_title = loan.book.title

        # Log the cancellation
        AuditLog.objects.create(
            action='loan_cancelled',
            user=request.user,
            target_user=request.user,
            book=loan.book,
            description=f"Loan request for '{book_title}' cancelled by {request.user.get_full_name()}",
            ip_address=request.META.get('REMOTE_ADDR')
        )

        loan.delete()

        messages.success(request, f'Your request for "{book_title}" has been cancelled.')
        return redirect('student_dashboard')

    context = {
        'loan': loan,
    }

    return render(request, 'library/cancel_request.html', context)


@login_required
def my_loans(request):
    """Student's loan management page"""
    # Get user's loans
    active_loans = Loan.objects.filter(
        borrower=request.user,
        status='active'
    ).select_related('book').order_by('-issue_date')

    overdue_loans = Loan.objects.filter(
        borrower=request.user,
        status='overdue'
    ).select_related('book').order_by('-issue_date')

    returned_loans = Loan.objects.filter(
        borrower=request.user,
        status='returned'
    ).select_related('book').order_by('-return_date')

    total_loans = Loan.objects.filter(
        borrower=request.user
    ).select_related('book').order_by('-issue_date')

    context = {
        'active_loans': active_loans,
        'overdue_loans': overdue_loans,
        'returned_loans': returned_loans,
        'total_loans': total_loans,
    }

    return render(request, 'library/my_loans.html', context)


@login_required
def my_reservations(request):
    """Student's reservations page"""
    reservations = Reservation.objects.filter(
        user=request.user
    ).select_related('book').order_by('-reservation_date')

    context = {
        'reservations': reservations,
    }

    return render(request, 'library/my_reservations.html', context)


@user_passes_test(is_librarian_or_admin)
def reservation_list(request):
    """Admin view for all reservations"""
    reservations = Reservation.objects.select_related('book', 'user').order_by('-reservation_date')

    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        reservations = reservations.filter(status=status_filter)

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        reservations = reservations.filter(
            Q(book__title__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(reservations, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'reservations': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'status_choices': Reservation.STATUS_CHOICES,
    }

    return render(request, 'library/reservation_list.html', context)


@user_passes_test(is_librarian_or_admin)
def overdue_loans(request):
    """View for overdue loans - books past due date"""
    from datetime import timedelta

    # Get overdue loans
    overdue_loans = Loan.objects.filter(
        status='active',
        due_date__lt=timezone.now()
    ).select_related('book', 'borrower').order_by('due_date')

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        overdue_loans = overdue_loans.filter(
            Q(book__title__icontains=search_query) |
            Q(borrower__username__icontains=search_query) |
            Q(borrower__first_name__icontains=search_query) |
            Q(borrower__last_name__icontains=search_query)
        )

    # Calculate days overdue for each loan
    for loan in overdue_loans:
        loan.days_overdue = (timezone.now().date() - loan.due_date.date()).days
        loan.fine_amount = loan.days_overdue * 1.00  # $1 per day fine

    # Pagination
    paginator = Paginator(overdue_loans, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'loans': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'page_title': 'Overdue Books',
        'page_description': 'Books that are past their due date and need immediate attention',
        'is_overdue': True,
    }

    return render(request, 'library/loan_tracking.html', context)


@user_passes_test(is_librarian_or_admin)
def due_soon_loans(request):
    """View for loans due soon - books due in next 3 days"""
    from datetime import timedelta

    # Get loans due in next 3 days
    today = timezone.now().date()
    due_soon_date = today + timedelta(days=3)

    due_soon_loans = Loan.objects.filter(
        status='active',
        due_date__date__gte=today,
        due_date__date__lte=due_soon_date
    ).select_related('book', 'borrower').order_by('due_date')

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        due_soon_loans = due_soon_loans.filter(
            Q(book__title__icontains=search_query) |
            Q(borrower__username__icontains=search_query) |
            Q(borrower__first_name__icontains=search_query) |
            Q(borrower__last_name__icontains=search_query)
        )

    # Calculate days until due for each loan
    for loan in due_soon_loans:
        loan.days_until_due = (loan.due_date.date() - today).days

    # Pagination
    paginator = Paginator(due_soon_loans, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'loans': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'page_title': 'Books Due Soon',
        'page_description': 'Books that are due within the next 3 days',
        'is_due_soon': True,
    }

    return render(request, 'library/loan_tracking.html', context)


@login_required
def my_pending_requests(request):
    """Student's pending requests page"""
    pending_requests = Loan.objects.filter(
        borrower=request.user,
        status='pending'
    ).select_related('book').order_by('-issue_date')

    context = {
        'pending_requests': pending_requests,
    }

    return render(request, 'library/my_pending_requests.html', context)


@login_required
def notifications_list(request):
    """Full notifications page"""
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')

    # Pagination
    paginator = Paginator(notifications, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'notifications': page_obj,
        'page_obj': page_obj,
    }

    return render(request, 'library/notifications_list.html', context)


@api_login_required
def api_notifications(request):
    """API endpoint for notifications"""
    from django.http import JsonResponse
    import traceback

    try:
        # Get base queryset for user notifications
        base_queryset = Notification.objects.filter(user=request.user).order_by('-created_at')

        # Get recent notifications (limited to 10)
        recent_notifications = base_queryset[:10]

        # Get unread count from base queryset (before slicing)
        unread_count = base_queryset.filter(is_read=False).count()

        notifications_data = []
        for notification in recent_notifications:
            notifications_data.append({
                'id': str(notification.id),
                'title': notification.title,
                'message': notification.message,
                'notification_type': notification.notification_type,
                'is_read': notification.is_read,
                'created_at': notification.created_at.isoformat(),
                'book_title': notification.book.title if notification.book else None,
            })

        return JsonResponse({
            'notifications': notifications_data,
            'unread_count': unread_count
        })
    except Exception as e:
        print(f"Error in api_notifications: {e}")
        print(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({
            'error': f'Failed to load notifications: {str(e)}',
            'notifications': [],
            'unread_count': 0
        }, status=500)


@api_login_required
def api_mark_notification_read(request, notification_id):
    """API endpoint to mark notification as read"""
    from django.http import JsonResponse

    if request.method == 'POST':
        try:
            notification = Notification.objects.get(id=notification_id, user=request.user)
            notification.is_read = True
            notification.save()
            return JsonResponse({'success': True})
        except Notification.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Notification not found'})

    return JsonResponse({'success': False, 'error': 'Invalid method'})


@api_login_required
def api_mark_all_notifications_read(request):
    """API endpoint to mark all notifications as read"""
    from django.http import JsonResponse

    if request.method == 'POST':
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Invalid method'})





def get_time_ago(datetime_obj):
    """Helper function to get human-readable time difference"""
    from django.utils import timezone
    from datetime import timedelta

    now = timezone.now()
    diff = now - datetime_obj

    if diff.days > 0:
        return f"{diff.days} day{'s' if diff.days > 1 else ''} ago"
    elif diff.seconds > 3600:
        hours = diff.seconds // 3600
        return f"{hours} hour{'s' if hours > 1 else ''} ago"
    elif diff.seconds > 60:
        minutes = diff.seconds // 60
        return f"{minutes} minute{'s' if minutes > 1 else ''} ago"
    else:
        return "Just now"


def create_notification(user, title, message, notification_type='general', book=None, loan=None):
    """Helper function to create notifications"""
    Notification.objects.create(
        user=user,
        title=title,
        message=message,
        notification_type=notification_type,
        book=book,
        loan=loan
    )


def send_email_notification(user, subject, message, send_to_additional_email=True):
    """Send email notification to user and optionally to their additional notification email"""
    from django.core.mail import send_mail
    from django.conf import settings

    recipients = [user.email]

    # Add additional notification email if provided and requested
    if send_to_additional_email and user.notification_email:
        recipients.append(user.notification_email)

    try:
        send_mail(
            subject=subject,
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=recipients,
            fail_silently=True,  # Don't break if email fails
        )
        return True
    except Exception as e:
        print(f"Failed to send email notification: {e}")
        return False


def send_parent_notification(student, subject, message, book_title=None):
    """Send email notification to student's parent/guardian"""
    from django.core.mail import send_mail
    from django.conf import settings

    # Only send to students who have parent email
    if student.role == 'student' and student.parent_email:
        try:
            # Customize message for parent
            parent_message = f"""
Dear {student.parent_name or 'Parent/Guardian'},

This is a notification regarding your ward {student.get_full_name()}'s library activity.

{message}

If you have any questions, please contact the school library.

Best regards,
School Library Team
            """.strip()

            send_mail(
                subject=f"Library Notification - {student.get_full_name()}",
                message=parent_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[student.parent_email],
                fail_silently=True,  # Don't break if email fails
            )
            return True
        except Exception as e:
            print(f"Failed to send parent notification: {e}")
            return False
    return False


def send_loan_notification(loan, notification_type):
    """Send notifications for loan events to both student and parent"""
    student = loan.borrower
    book = loan.book

    # Define notification content based on type
    notifications = {
        'loan_approved': {
            'title': 'Book Loan Approved',
            'student_message': f'Your request for "{book.title}" has been approved. Please visit the library to collect your book.',
            'parent_message': f'Your ward has been approved to borrow the book "{book.title}". Due date: {loan.due_date.strftime("%B %d, %Y")}.',
        },
        'book_due_soon': {
            'title': 'Book Due Soon',
            'student_message': f'Your book "{book.title}" is due on {loan.due_date.strftime("%B %d, %Y")}. Please return it on time to avoid fines.',
            'parent_message': f'Your ward\'s book "{book.title}" is due on {loan.due_date.strftime("%B %d, %Y")}. Please remind them to return it on time.',
        },
        'book_overdue': {
            'title': 'Book Overdue',
            'student_message': f'Your book "{book.title}" is overdue. Please return it immediately to avoid additional fines.',
            'parent_message': f'Your ward\'s book "{book.title}" is overdue. Please ensure they return it immediately to avoid additional fines.',
        },
        'book_returned': {
            'title': 'Book Returned',
            'student_message': f'Thank you for returning "{book.title}". Your loan has been completed successfully.',
            'parent_message': f'Your ward has successfully returned the book "{book.title}". Thank you for ensuring timely return.',
        }
    }

    if notification_type in notifications:
        notif = notifications[notification_type]

        # Create in-app notification for student
        create_notification(
            user=student,
            title=notif['title'],
            message=notif['student_message'],
            notification_type=notification_type.replace('_', '_'),
            book=book,
            loan=loan
        )

        # Send email to student
        try:
            from django.core.mail import send_mail
            from django.conf import settings

            send_mail(
                subject=notif['title'],
                message=notif['student_message'],
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[student.email],
                fail_silently=True,
            )
        except:
            pass

        # Send email to parent if student
        send_parent_notification(
            student=student,
            subject=notif['title'],
            message=notif['parent_message'],
            book_title=book.title
        )
